from io import BytesIO
from datetime import date, timedelta
from calendar import monthrange
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from .db import Person, Audit
from .rules import summarize

def safe(value):
    return "'"+value if isinstance(value,str) and value.startswith(('=','+','-','@')) else value
def export_month(db,month,closing):
    year,number=map(int,month.split('-')); rows=[]
    for p in db.scalars(select(Person).order_by(Person.name)):
        for n in range(1,monthrange(year,number)[1]+1):
            d=date(year,number,n).isoformat()
            if p.hired<=d and (not p.terminated or d<=p.terminated): rows.append(summarize(db,p,d))
    count=max([r['expected'] for r in rows]+[len(r['punches']) for r in rows]+[2])
    wb=Workbook(); ws=wb.active; ws.title='Pontos - Geral'
    ws.append(['Data','ID','Pessoa','Perfil']+[f'Batida {i+1}' for i in range(count)]+['Jornada vigente','Nº batidas','Previsto (min)','Trabalhado (min)','Atraso (min)','Extra (min)','Negativo (min)','Saldo (min)','Situação','Validação da extra','Teste','Desligamento'])
    monthly=defaultdict(lambda:[0]*9); weekly=defaultdict(lambda:[0]*9)
    for r in rows:
        times=[p['time'] for p in r['punches']]
        ws.append([date.fromisoformat(r['date']),r['person_id'],safe(r['name']),r['role']]+times+['']*(count-len(times))+[' / '.join('–'.join(x) for x in r['periods']),len(times),r['planned'],r['worked'],r['late'],r['extra'],r['negative'],r['balance'],safe(r['status']),r['overtime'],'Sim' if any(p.get('test') for p in r['punches']) else '',r['terminated']])
        ws.cell(ws.max_row,1).number_format='dd/mm/yyyy'
        if r['absent'] and not times and not r['holiday']:
            ws.merge_cells(start_row=ws.max_row,start_column=5,end_row=ws.max_row,end_column=4+max(2,r['expected']))
            ws.cell(ws.max_row,5,'FALTA REGISTRADA')
        if r['absent'] or 'incompleto' in r['status'] or (r['expected'] and not times and not r['holiday']):
            for c in ws[ws.max_row]: c.font=Font(color='B42318'); c.fill=PatternFill('solid',fgColor='FFF1F0')
        day=date.fromisoformat(r['date']); week=(day-timedelta(days=day.weekday())).isoformat()
        values=[r['planned'],r['worked'],r['late'],r['extra'],r['negative'],r['balance'] or 0,int(bool(times)),int(r['balance'] is None and bool(r['expected'])),int(r['absent'])]
        for collection,key in [(monthly,(r['person_id'],r['name'])),(weekly,(r['person_id'],r['name'],week))]: collection[key]=[a+b for a,b in zip(collection[key],values)]
    for title,collection in [('Resumo por funcionário',monthly),('Resumo semanal',weekly)]:
        sheet=wb.create_sheet(title)
        sheet.append(['ID','Pessoa']+(['Semana iniciada em'] if collection is weekly else [])+['Previsto (min)','Trabalhado (min)','Atraso (min)','Extra (min)','Negativo (min)','Saldo apurado (min)','Dias com batida','Dias sem apuração completa','Faltas registradas'])
        for key,values in collection.items(): sheet.append([safe(v) for v in key]+values)
    sheet=wb.create_sheet('Alterações')
    sheet.append(['Quando','Responsável','Ação','Referência','Anterior','Novo','Motivo'])
    import json
    for a in db.scalars(select(Audit).order_by(Audit.id)):
        if month in a.target or a.at.startswith(month): sheet.append([a.at,safe(a.actor),a.action,a.target,safe(json.dumps(a.before,ensure_ascii=False)),safe(json.dumps(a.after,ensure_ascii=False)),safe(a.reason)])
    wb.properties.title=f'Livro-Ponto {month} — versão {closing["version"]}'
    wb.properties.description=f'Fechado por {closing["by"]} em {closing["at"]}. Minutos inteiros. Saldo exclui dias sem apuração completa. Extras dependem de validação.'
    for sheet in wb:
        sheet.freeze_panes='E2' if sheet==ws else 'C2'; sheet.auto_filter.ref=sheet.dimensions
        sheet.sheet_view.showGridLines=False
        sheet.print_title_rows='1:1'; sheet.page_setup.orientation='landscape'; sheet.page_setup.paperSize=sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth=1; sheet.sheet_properties.pageSetUpPr.fitToPage=True
        sheet.oddFooter.center.text=f'{month} • versão {closing["version"]} • &P / &N'
        sheet.row_dimensions[1].height=42
        for cell in sheet[1]:
            cell.font=Font(name='Calibri',size=11,bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='18334F'); cell.alignment=Alignment(wrap_text=True,vertical='center')
        for idx,column in enumerate(sheet.columns,1):
            sheet.column_dimensions[get_column_letter(idx)].width=min(48,max(15,max(len(str(c.value or '')) for c in column)+2))
        for row in sheet.iter_rows(min_row=2):
            for c in row:
                c.alignment=Alignment(vertical='center')
                if c.fill.patternType is None and c.row%2==0: c.fill=PatternFill('solid',fgColor='F1F5F9')
    output=BytesIO(); wb.save(output); return output.getvalue()
